import json
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def get_info(path):
    with open(path, 'r', encoding='utf-8')\
     as f:
        all_info = []
        json_data = f.read()
        dict_data = json.loads(json_data)
        # 表名称
        table_name = dict_data.get('tableName')
        # 表备注
        table_comment = dict_data.get('displayName')
        tables_comment_dict = json.loads(dict_data['dataDesc'])
        tmp = [table_name, table_comment]
        all_info.append(tmp)
        for a_colum in tables_comment_dict.keys():
            colum_comment = tables_comment_dict.get(a_colum).get('title')
            cell_info = [a_colum, colum_comment]
            all_info.append(cell_info)
        return all_info


def get_all_files(root_path):
    """
    获得全部json文件的路径
    root_path: json文件所在的路径
    return all_files: 所有的文件的全路径
    """
    all_files = []
    tmp_list = []
    for root, dirs, files in os.walk(root_path):
        for a_file in files:
            tmp_list.append(os.path.join(root + '/', a_file))
        all_files = [x for x in tmp_list if x.endswith('.mod')]
    return all_files


def get_all_excel(file_path):
    """
    获得全部的excel文件
    :param file_path:
    :return:
    """
    for root, dirs, files in os.walk(root_path):
        all_excel = [x for x in files]
    return all_excel


def mapping_comment(file_path, data_info):
    """
    匹配表中的字段与备注
    :param file_path: 需要匹配的文件
    :param data_info: 传入的字段信息
    :return:
    """
    wb = load_workbook(file_path)
    ws = wb.active
    # 插入空白行，留来写入字段备注
    ws.insert_rows(2)
    row1 = ws[1]
    # 表头存储列表 存储了表头字段和索引
    index_cell_tmp = []
    # 获得表字段和对应索引
    for k, v in enumerate(row1):
        tmp = [v.value, k + 1]
        index_cell_tmp.append(tmp)
    # 去掉表名和表备注
    data_info = data_info[1:]
    for cell_comment in data_info:
        for index_cell in index_cell_tmp:
            if cell_comment[0] == index_cell[0]:
                # 列号
                column_index = get_column_letter(index_cell[1])
                # 写入字段备注
                ws[str(column_index) + '2'] = cell_comment[1]
                # print(get_column_letter(index_cell[1]))
    wb.save(file_path)


if __name__ == '__main__':
    all_file_info = []
    all_json = get_all_files('C:/Users/nianfou/Desktop/module')
    for a_json in all_json:
        file_info = get_info(a_json)
        all_file_info.append(file_info)
    for a_info in all_file_info:
        root_path = 'D:/案件/2020-0141/刻盘数据'
        file_path = os.path.join(root_path+'/', a_info[0][0]+'.xlsx')
        if os.path.exists(file_path):
            mapping_comment(file_path, a_info)
            old_path = file_path
            new_path = os.path.join(root_path+'/', a_info[0][0]+a_info[0][1]+'.xlsx')
            try:
                os.rename(old_path, new_path)
            except FileNotFoundError:
                print('{}重命名失败，请重新命名'.format(a_info[0][0]))
        else:
            print("{}文件不存在,请手动检查".format(a_info[0][0]))

